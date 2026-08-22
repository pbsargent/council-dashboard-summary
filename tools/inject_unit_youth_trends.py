#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def clean_header(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).replace("\n", " ").split()).strip()
    return text or None


def num(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.strip().startswith("#"):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return int(parsed) if parsed.is_integer() else parsed


def month_values(ws: Any, row: int) -> list[float | int | None]:
    return [num(clean_value(ws.cell(row, col).value)) for col in range(2, 14)]


def latest_filled_month(months: list[str], values: list[Any]) -> str | None:
    latest = None
    for index, value in enumerate(values):
        if value is not None:
            latest = months[index]
    return latest


def build_unit_youth_trends(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Units-Youth" not in wb.sheetnames:
        raise RuntimeError(f"{workbook_path.name} is missing required sheet: Units-Youth")

    ws = wb["Units-Youth"]
    months = [
        clean_header(ws.cell(2, col).value)
        or clean_header(ws.cell(10, col).value)
        or clean_header(ws.cell(23, col).value)
        or f"Month {col - 1}"
        for col in range(2, 14)
    ]

    series = {
        "new_units": {
            "label": "New Units",
            "values": {
                "2026": month_values(ws, 3),
                "2025": month_values(ws, 4),
                "2024": month_values(ws, 6),
            },
            "delta_vs_prior_year": month_values(ws, 5),
        },
        "new_youth": {
            "label": "New Youth",
            "values": {
                "2026": month_values(ws, 11),
                "2025": month_values(ws, 12),
                "2024": month_values(ws, 14),
            },
            "delta_vs_prior_year": month_values(ws, 13),
        },
        "total_youth": {
            "label": "Total Youth",
            "values": {
                "2026": month_values(ws, 24),
                "2025": month_values(ws, 25),
                "2024": month_values(ws, 26),
            },
        },
        "total_units": {
            "label": "Total Units",
            "values": {
                "2026": month_values(ws, 31),
                "2025": month_values(ws, 32),
                "2024": month_values(ws, 33),
            },
        },
    }

    for item in series.values():
        item["freshness_month"] = latest_filled_month(months, item["values"]["2026"])

    return {
        "source_sheet": "Units-Youth",
        "source_note": "Manual workbook tab. Freshness can vary by block, so each series carries its own latest filled month.",
        "months": months,
        "ytd": {
            "current_period": clean_value(ws.cell(19, 1).value),
            "prior_period": clean_value(ws.cell(20, 1).value),
            "new_units": num(ws.cell(19, 2).value),
            "prior_new_units": num(ws.cell(20, 2).value),
            "new_units_delta": num(ws.cell(21, 2).value),
            "new_youth": num(ws.cell(19, 3).value),
            "prior_new_youth": num(ws.cell(20, 3).value),
            "new_youth_delta": num(ws.cell(21, 3).value),
        },
        "series": series,
    }


def inject(json_path: Path) -> dict[str, Any]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    workbook_path = Path(payload["dashboard"]["source"])
    payload["dashboard"]["unit_youth_trends"] = build_unit_youth_trends(workbook_path)
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return payload["dashboard"]["unit_youth_trends"]


def main() -> int:
    parser = argparse.ArgumentParser(description="Inject Units-Youth workbook trends into dashboard JSON.")
    parser.add_argument("json_paths", nargs="+", type=Path)
    args = parser.parse_args()

    for json_path in args.json_paths:
        trends = inject(json_path)
        print(
            json.dumps(
                {
                    "json": str(json_path),
                    "source_sheet": trends["source_sheet"],
                    "series": sorted(trends["series"].keys()),
                },
                indent=2,
            )
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
