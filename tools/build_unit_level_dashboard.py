#!/usr/bin/env python3
"""Build the Unit Level Dashboard JSON from the Unit Level Metrics workbook.

This follows the Council dashboard's source pattern: read the locally available
workbook with openpyxl/data_only, publish a compact JSON snapshot, and keep the
browser independent from Excel and OneDrive.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return None
    return value


def yes(value):
    return str(value or "").strip().lower() == "yes"


def records(sheet, header_row=1):
    headers = [cell.value for cell in sheet[header_row]]
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        yield {str(headers[index]).strip(): value for index, value in enumerate(values) if index < len(headers) and headers[index]}


def unit_name(row):
    name = str(row.get("Name") or "").strip()
    if name:
        return name
    parts = [str(row.get("Unit") or "").strip(), str(row.get("#") or "").strip(), str(row.get("Gender") or "").strip()]
    return " ".join(part for part in parts if part)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("data/unit-level-latest.json"))
    parser.add_argument("--js-output", type=Path, default=Path("data/unit-level-latest.js"))
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    required = {"Unit_Metrics", "Units", "MembersDueToRenew"}
    missing = sorted(required.difference(workbook.sheetnames))
    if missing:
        raise RuntimeError(f"Workbook is missing required sheets: {', '.join(missing)}")

    training_by_id = {}
    for row in records(workbook["Units"]):
        unit_id = row.get("UnitID")
        if unit_id in (None, ""):
            continue
        training_by_id[int(unit_id)] = {
            "direct_contact_trained_rate": clean(row.get("Direct Contact Trained")),
            "all_leaders_trained_rate": clean(row.get("All Ldr Trained")),
            "baloo_trained": clean(row.get("BALOO Trained")),
            "syt_compliance_rate": clean(row.get("SYT")),
            "syt_0_30": clean(row.get("SYT 0-30")) or 0,
            "syt_31_90": clean(row.get("SYT 31-90")) or 0,
        }

    renewals = defaultdict(lambda: {"records": 0, "youth": 0, "adults": 0, "opt_outs": 0, "dates": defaultdict(int), "members": []})
    for row in records(workbook["MembersDueToRenew"]):
        unit_id = row.get("UnitID")
        if not isinstance(unit_id, (int, float)):
            continue
        summary = renewals[int(unit_id)]
        summary["records"] += 1
        if str(row.get("Youth") or "").strip().lower() == "yes":
            summary["youth"] += 1
        else:
            summary["adults"] += 1
        if str(row.get("Opt_Out") or "").strip().lower() == "yes":
            summary["opt_outs"] += 1
        renewal = row.get("Renewal")
        renewal_date = None
        if isinstance(renewal, (datetime, date)):
            renewal_date = renewal.date().isoformat() if isinstance(renewal, datetime) else renewal.isoformat()
            summary["dates"][renewal_date] += 1
        summary["members"].append({
            "name": clean(row.get("Name")),
            "renewal_date": renewal_date,
            "type": "Youth" if str(row.get("Youth") or "").strip().lower() == "yes" else "Adult",
            "opt_out": str(row.get("Opt_Out") or "").strip().lower() == "yes",
        })

    units = []
    for row in records(workbook["Unit_Metrics"]):
        unit_id = row.get("Unit_ID")
        metric = row.get("Unit Metric")
        if not isinstance(unit_id, (int, float)) or not isinstance(metric, (int, float)):
            continue
        unit_id = int(unit_id)
        youth = row.get("Youth") if isinstance(row.get("Youth"), (int, float)) else None
        prior = row.get("Youth Prev Yr") if isinstance(row.get("Youth Prev Yr"), (int, float)) else None
        due = renewals.get(unit_id, {"records": 0, "youth": 0, "adults": 0, "opt_outs": 0, "dates": {}, "members": []})
        units.append({
            "unit_id": unit_id,
            "district": clean(row.get("District")),
            "name": unit_name(row),
            "unit_type": clean(row.get("Unit")),
            "number": clean(row.get("#")),
            "gender": clean(row.get("Gender")),
            "metric": metric,
            "youth": youth,
            "youth_prior": prior,
            "youth_change": youth - prior if youth is not None and prior is not None else None,
            "retention_pct": clean(row.get("Retention")),
            "advancement_pct": clean(row.get("Advancement")),
            "last_outdoor_date": clean(row.get("Super Activity")),
            "last_connection": clean(row.get("Connection")),
            "commissioner": clean(row.get("Commissioner")),
            "chartered_organization": clean(row.get("Chartered Organization")),
            "drivers": {
                "ul_cc_trained": yes(row.get("UL-CC Trained")),
                "size": yes(row.get("Exceed Small Unit Threshold")),
                "growth": yes(row.get("YOY Growth")),
                "advancement_or_officers": yes(row.get("Advancement OR Officers")),
                "outdoor": yes(row.get("Outdoor")),
            },
            "unit_leader_trained": yes(row.get("UL Trained")),
            "committee_chair_trained": yes(row.get("CC Trained")),
            "training": training_by_id.get(unit_id, {}),
            "renewal_records": {
                "total": due["records"],
                "youth": due["youth"],
                "adults": due["adults"],
                "opt_outs": due["opt_outs"],
                "by_date": dict(sorted(due["dates"].items())),
                "members": sorted(due["members"], key=lambda member: (member["renewal_date"] or "", member["name"] or "")),
            },
        })

    units.sort(key=lambda row: (row["district"] or "", row["name"] or ""))
    snapshot_date = datetime.fromtimestamp(args.workbook.stat().st_mtime).date().isoformat()
    output = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "data_date": snapshot_date,
        "source": {
            "name": args.workbook.name,
            "method": "openpyxl read_only=True, data_only=True; compact JSON snapshot",
            "sheets": ["Unit_Metrics", "Units", "MembersDueToRenew"],
        },
        "default_unit_id": 258381 if any(row["unit_id"] == 258381 for row in units) else (units[0]["unit_id"] if units else None),
        "units": units,
    }

    payload = json.dumps(output, indent=2, ensure_ascii=False)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(payload, encoding="utf-8")
    args.js_output.parent.mkdir(parents=True, exist_ok=True)
    args.js_output.write_text(f"window.UNIT_LEVEL_DATA = {payload};\n", encoding="utf-8")
    print(f"Wrote {len(units)} units to {args.output} and {args.js_output}")


if __name__ == "__main__":
    main()
