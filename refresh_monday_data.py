#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


API_URL = "https://api.monday.com/v2"
DEFAULT_TOKEN_FILES = [
    Path("/Users/petersargent/Documents/06 Personal, Legal, and Sensitive/Sensitive - Move to Password Manager/Monday-Com-API-Token.txt"),
    Path("/Users/petersargent/Documents/Monday-Com-API-Token.txt"),
]
DEFAULT_TOKEN_FILE = DEFAULT_TOKEN_FILES[0]
DEFAULT_WORKBOOK_DIR = Path(
    "/Users/petersargent/Library/CloudStorage/GoogleDrive-peter@imetpetersargent.com/"
    "Shared drives/Council monday.com Reports"
)

BOARDS = {
    "prospects": {
        "id": 8502314751,
        "name": "New unit Hot Prospects",
        "url": "https://capitolareacouncil564.monday.com/boards/8502314751",
        "columns": {
            "district": "label__1",
            "projected_start": "color_mknw6473",
            "status": "status",
        },
    },
    "renewals": {
        "id": 18289783899,
        "name": "2026 Unit Renewal",
        "url": "https://capitolareacouncil564.monday.com/boards/18289783899",
        "columns": {
            "intent": "color_mkx5j65h",
            "posted": "color_mkx59sx7",
        },
    },
    "schools": {
        "id": 9158891570,
        "name": "Schools",
        "url": "https://capitolareacouncil564.monday.com/boards/9158891570",
        "columns": {
            "status": "dropdown_mm3gpfex",
            "district": "dropdown_mkqzfzs3",
        },
    },
}


def read_token(path: Path) -> str:
    candidates = [path]
    candidates.extend(p for p in DEFAULT_TOKEN_FILES if p != path)
    token_path = next((candidate for candidate in candidates if candidate.exists()), None)
    if token_path is None:
        checked = ", ".join(str(candidate) for candidate in candidates)
        raise FileNotFoundError(f"Missing monday.com API token file. Checked: {checked}")
    for line in token_path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text or text.startswith("#"):
            continue
        if "=" in text:
            text = text.split("=", 1)[1].strip()
        return text.strip("\"'")
    raise ValueError(f"No API token found in {token_path}")


def monday_query(token: str, query: str, variables: dict[str, Any]) -> dict[str, Any]:
    payload = json.dumps({"query": query, "variables": variables}).encode("utf-8")
    request = Request(
        API_URL,
        data=payload,
        headers={
            "Authorization": token,
            "Content-Type": "application/json",
            "API-Version": "2025-04",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=45) as response:
            body = response.read().decode("utf-8")
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"monday.com API HTTP {error.code}: {detail}") from error
    except URLError as error:
        raise RuntimeError(f"monday.com API connection failed: {error}") from error

    parsed = json.loads(body)
    if parsed.get("errors"):
        raise RuntimeError(f"monday.com API returned errors: {parsed['errors']}")
    return parsed["data"]


def fetch_board_items(token: str, board_id: int, column_ids: list[str]) -> list[dict[str, Any]]:
    query = """
    query FetchBoardItems($boardId: [ID!], $columnIds: [String!], $cursor: String) {
      boards(ids: $boardId) {
        items_page(limit: 500, cursor: $cursor) {
          cursor
          items {
            id
            name
            updated_at
            column_values(ids: $columnIds) {
              id
              text
              value
            }
          }
        }
      }
    }
    """
    items: list[dict[str, Any]] = []
    cursor = None
    while True:
        data = monday_query(token, query, {"boardId": [str(board_id)], "columnIds": column_ids, "cursor": cursor})
        boards = data.get("boards") or []
        if not boards:
            raise RuntimeError(f"Board {board_id} was not returned by monday.com")
        page = boards[0]["items_page"]
        items.extend(page.get("items") or [])
        cursor = page.get("cursor")
        if not cursor:
            break
        time.sleep(0.15)
    return items


def column_map(item: dict[str, Any]) -> dict[str, str]:
    return {
        value.get("id"): (value.get("text") or "").strip()
        for value in item.get("column_values", [])
    }


def count_labels(
    items: list[dict[str, Any]],
    column_id: str,
    blank_label: str = "Unlabeled",
    split_multi: bool = False,
) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    for item in items:
        text = column_map(item).get(column_id) or ""
        labels = [label.strip() for label in text.split(",")] if split_multi and text else [text]
        labels = [label for label in labels if label]
        if not labels:
            labels = [blank_label]
        for label in labels:
            counts[label] += 1
    return sorted(
        [{"label": label, "count": count} for label, count in counts.items()],
        key=lambda row: (-row["count"], row["label"]),
    )


def latest_update(items: list[dict[str, Any]]) -> str | None:
    values = [item.get("updated_at") for item in items if item.get("updated_at")]
    return max(values) if values else None


def latest_workbook(source_dir: Path) -> Path | None:
    files = sorted(
        source_dir.glob("*monday-export.xlsx"),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    return files[0] if files else None


def as_text(value: Any) -> str:
    return str(value or "").strip()


def excel_timestamp(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        dt = from_excel(value)
    else:
        text = as_text(value)
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt.isoformat(timespec="seconds").replace("+00:00", "Z")


def workbook_generated_at(workbook: Any) -> str:
    if "Overview" not in workbook.sheetnames:
        return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    ws = workbook["Overview"]
    value = ws.cell(row=2, column=1).value
    text = as_text(value)
    if text.startswith("Extracted:"):
        parsed = excel_timestamp(text.split(":", 1)[1].strip())
        if parsed:
            return parsed
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def workbook_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Workbook is missing required sheet: {sheet_name}")
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise RuntimeError(f"Workbook sheet has no header row: {sheet_name}")
    headers = [as_text(value) for value in rows[3]]
    return [
        dict(zip(headers, row))
        for row in rows[4:]
        if any(value is not None for value in row)
    ]


def count_workbook_labels(
    rows: list[dict[str, Any]],
    column: str,
    blank_label: str = "Unlabeled",
    split_multi: bool = False,
) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    for row in rows:
        text = as_text(row.get(column))
        labels = [label.strip() for label in text.split(",")] if split_multi and text else [text]
        labels = [label for label in labels if label]
        if not labels:
            labels = [blank_label]
        for label in labels:
            counts[label] += 1
    return sorted(
        [{"label": label, "count": count} for label, count in counts.items()],
        key=lambda row: (-row["count"], row["label"]),
    )


def latest_workbook_update(rows: list[dict[str, Any]]) -> str | None:
    values = [excel_timestamp(row.get("Updated At")) for row in rows]
    values = [value for value in values if value]
    return max(values) if values else None


def compact_prospect_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{
        "item_id": as_text(row.get("Item ID")),
        "name": as_text(row.get("Item Name")),
        "group": as_text(row.get("Group")),
        "district": as_text(row.get("District")) or "Unassigned",
        "unit_type": as_text(row.get("Potential Unit Type(s)")),
        "unit_numbers": as_text(row.get("Unit Number(s)")),
        "projected_start": as_text(row.get("Projected Start Month")) or "Unscheduled",
        "status": as_text(row.get("Step 1")) or "Unlabeled",
        "contact_stage": as_text(row.get("Step 2")),
        "first_visit": as_text(row.get("Step 3")),
        "charter_stage": as_text(row.get("Step 10")),
        "posted": as_text(row.get("ONLY CLICK COMPLETE WHEN unit posts!!Unit in my.scouting?")),
        "first_meeting": as_text(row.get("Date of First Meeting")),
        "updated_at": excel_timestamp(row.get("Updated At")),
    } for row in rows]


def compact_renewal_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{
        "item_id": as_text(row.get("Item ID")),
        "name": as_text(row.get("Item Name")),
        "group": as_text(row.get("Group")),
        "district": as_text(row.get("District")) or "Unassigned",
        "intent": as_text(row.get("Drop/Renew")) or "Unlabeled",
        "initiated": as_text(row.get("Intiated")),
        "submitted": as_text(row.get("Submitted")),
        "pending_acceptance": as_text(row.get("Pending Acceptance")),
        "posted": as_text(row.get("Posted")) or "Not posted",
        "timeline": as_text(row.get("Timeline")),
        "updated_at": excel_timestamp(row.get("Updated At")),
    } for row in rows]


def compact_school_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{
        "item_id": as_text(row.get("Item ID")),
        "name": as_text(row.get("Item Name")),
        "group": as_text(row.get("Group")),
        "school_district": as_text(row.get("School District")) or "Unassigned",
        "scouting_district": as_text(row.get("Scouting District")) or "Unassigned",
        "unit_associated": as_text(row.get("Unit Associated")),
        "tay": as_text(row.get("TAY")),
        "principal_meeting": as_text(row.get("Principal Meeting")),
        "city": as_text(row.get("City")),
        "county": as_text(row.get("County Name")),
        "district_type": as_text(row.get("District Type")),
        "instruction_type": as_text(row.get("Instruction Type")),
        "status": as_text(row.get("School Status")) or "Unlabeled",
        "updated_at": excel_timestamp(row.get("Updated At")),
    } for row in rows]


def build_snapshot_from_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    prospects_rows = workbook_rows(workbook, "New unit Hot Prospects")
    renewals_rows = workbook_rows(workbook, "2026 Unit Renewal")
    schools_rows = workbook_rows(workbook, "Schools")

    return {
        "generated_from": "monday.com daily workbook",
        "generated_at": workbook_generated_at(workbook),
        "source_workbook": path.name,
        "boards": {
            "prospects": {
                "name": BOARDS["prospects"]["name"],
                "url": BOARDS["prospects"]["url"],
                "updated_at": latest_workbook_update(prospects_rows),
                "items": len(prospects_rows),
                "status": count_workbook_labels(prospects_rows, "Step 1"),
                "districts": count_workbook_labels(prospects_rows, "District", "Unassigned"),
                "projected_start_months": count_workbook_labels(prospects_rows, "Projected Start Month", "Unscheduled"),
                "rows": compact_prospect_rows(prospects_rows),
            },
            "renewals": {
                "name": BOARDS["renewals"]["name"],
                "url": BOARDS["renewals"]["url"],
                "updated_at": latest_workbook_update(renewals_rows),
                "items": len(renewals_rows),
                "intent": count_workbook_labels(renewals_rows, "Drop/Renew"),
                "posted": count_workbook_labels(renewals_rows, "Posted", "Not posted"),
                "rows": compact_renewal_rows(renewals_rows),
            },
            "schools": {
                "name": BOARDS["schools"]["name"],
                "url": BOARDS["schools"]["url"],
                "updated_at": latest_workbook_update(schools_rows),
                "items": len(schools_rows),
                "status": count_workbook_labels(schools_rows, "School Status"),
                "districts": count_workbook_labels(schools_rows, "Scouting District", "Unassigned", split_multi=True),
                "rows": compact_school_rows(schools_rows),
            },
        },
    }


def build_snapshot(token: str) -> dict[str, Any]:
    prospects_items = fetch_board_items(token, BOARDS["prospects"]["id"], list(BOARDS["prospects"]["columns"].values()))
    renewals_items = fetch_board_items(token, BOARDS["renewals"]["id"], list(BOARDS["renewals"]["columns"].values()))
    schools_items = fetch_board_items(token, BOARDS["schools"]["id"], list(BOARDS["schools"]["columns"].values()))

    prospects_columns = BOARDS["prospects"]["columns"]
    renewals_columns = BOARDS["renewals"]["columns"]
    schools_columns = BOARDS["schools"]["columns"]

    return {
        "generated_from": "monday.com API",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "boards": {
            "prospects": {
                "name": BOARDS["prospects"]["name"],
                "url": BOARDS["prospects"]["url"],
                "updated_at": latest_update(prospects_items),
                "items": len(prospects_items),
                "status": count_labels(prospects_items, prospects_columns["status"]),
                "districts": count_labels(prospects_items, prospects_columns["district"], "Unassigned"),
                "projected_start_months": count_labels(prospects_items, prospects_columns["projected_start"], "Unscheduled"),
            },
            "renewals": {
                "name": BOARDS["renewals"]["name"],
                "url": BOARDS["renewals"]["url"],
                "updated_at": latest_update(renewals_items),
                "items": len(renewals_items),
                "intent": count_labels(renewals_items, renewals_columns["intent"]),
                "posted": count_labels(renewals_items, renewals_columns["posted"], "Not posted"),
            },
            "schools": {
                "name": BOARDS["schools"]["name"],
                "url": BOARDS["schools"]["url"],
                "updated_at": latest_update(schools_items),
                "items": len(schools_items),
                "status": count_labels(schools_items, schools_columns["status"]),
                "districts": count_labels(schools_items, schools_columns["district"], "Unassigned", split_multi=True),
            },
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Refresh Council Dashboard Summary monday.com data.")
    parser.add_argument("--token-file", type=Path, default=DEFAULT_TOKEN_FILE)
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_WORKBOOK_DIR)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    workbook_path = args.workbook or latest_workbook(args.source_dir)
    if workbook_path:
        try:
            snapshot = build_snapshot_from_workbook(workbook_path)
        except Exception as error:
            print(f"Workbook refresh failed ({error}); falling back to monday.com API.", file=sys.stderr)
            token = read_token(args.token_file)
            snapshot = build_snapshot(token)
    else:
        print(f"No monday.com export workbook found in {args.source_dir}; falling back to monday.com API.", file=sys.stderr)
        token = read_token(args.token_file)
        snapshot = build_snapshot(token)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "generated_from": snapshot["generated_from"],
        "prospects": snapshot["boards"]["prospects"]["items"],
        "renewals": snapshot["boards"]["renewals"]["items"],
        "schools": snapshot["boards"]["schools"]["items"],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
